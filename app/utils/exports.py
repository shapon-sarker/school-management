from weasyprint import HTML, CSS
from openpyxl import Workbook
from datetime import datetime
import os

class ExportManager:
    @staticmethod
    def export_to_pdf(template_html, css_file=None, output_file=None):
        """
        Export HTML template to PDF
        """
        if css_file:
            css = CSS(filename=css_file)
        else:
            css = None
        
        html = HTML(string=template_html)
        if not output_file:
            output_file = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        html.write_pdf(output_file, stylesheets=[css] if css else None)
        return output_file

    @staticmethod
    def export_to_excel(data, headers, output_file=None):
        """
        Export data to Excel
        """
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        if not output_file:
            output_file = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(output_file)
        return output_file

class IDCardGenerator:
    @staticmethod
    def generate_id_card(student_data, template_path, output_file=None):
        """
        Generate student ID card using HTML template
        """
        # Load the ID card template
        with open(template_path, 'r') as f:
            template = f.read()
        
        # Replace placeholders with student data
        for key, value in student_data.items():
            template = template.replace(f'{{{{ {key} }}}}', str(value))
        
        # Generate PDF
        if not output_file:
            output_file = f'id_card_{student_data["admission_number"]}.pdf'
        
        return ExportManager.export_to_pdf(template, output_file=output_file) 